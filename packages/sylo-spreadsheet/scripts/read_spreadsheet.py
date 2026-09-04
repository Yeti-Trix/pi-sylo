#!/usr/bin/env python3
"""Read .xlsx or .ods workbook sheet as JSON (bounded rows for agent context)."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


SUPPORTED_EXTENSIONS = {".xlsx", ".ods"}
DEFAULT_MAX_ROWS = 200


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {letters!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _parse_a1_range(spec: str) -> tuple[int, int, int, int]:
    """Return min_row, max_row, min_col, max_col (1-based)."""
    cleaned = spec.strip().upper()
    m = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cleaned)
    if not m:
        raise ValueError(f"Invalid range {spec!r}; use A1 notation like B2:F50")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    min_col, max_col = _col_letters_to_index(c1), _col_letters_to_index(c2)
    min_row, max_row = r1, r2
    if min_row > max_row or min_col > max_col:
        raise ValueError(f"Invalid range {spec!r}; corners must form a rectangle")
    return min_row, max_row, min_col, max_col


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


def _resolve_sheet_name(sheet_names: list[str], sheet_arg: str | None) -> str:
    if not sheet_names:
        raise ValueError("Workbook has no sheets")
    if not sheet_arg:
        return sheet_names[0]
    if sheet_arg.isdigit():
        idx = int(sheet_arg)
        if idx < 1 or idx > len(sheet_names):
            raise ValueError(f"Sheet index {idx} out of range (1–{len(sheet_names)})")
        return sheet_names[idx - 1]
    if sheet_arg not in sheet_names:
        raise ValueError(f"Sheet {sheet_arg!r} not found; available: {', '.join(sheet_names)}")
    return sheet_arg


def _read_xlsx(
    path: Path,
    sheet_name: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    max_rows: int,
    include_formulas: bool,
) -> tuple[list[list[Any]], dict[str, int], bool]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=not include_formulas)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found")
        ws = wb[sheet_name]
        total_rows_in_window = max(0, max_row - min_row + 1)
        data_row_budget = max_rows if max_rows > 0 else total_rows_in_window
        # First row in window is headers; cap data rows after that.
        allowed_rows = min(total_rows_in_window, 1 + data_row_budget)
        effective_max_row = min_row + allowed_rows - 1

        matrix: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=effective_max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            matrix.append([_normalize_cell(v) for v in row])

        truncated = effective_max_row < max_row or (max_rows > 0 and total_rows_in_window > allowed_rows)
        dimensions = {
            "total_rows_in_range": total_rows_in_window,
            "returned_rows": len(matrix),
            "col_count": max_col - min_col + 1,
        }
        return matrix, dimensions, truncated
    finally:
        wb.close()


def _ods_cell_text(cell) -> Any:
    from odf.text import P

    parts = [str(p) for p in cell.getElementsByType(P)]
    value = "".join(parts).strip()
    if value == "":
        return None
    if re.fullmatch(r"-?\d+", value):
        return int(value)
    if re.fullmatch(r"-?\d+\.\d+", value):
        return float(value)
    return value


def _expand_ods_row(row, min_col: int, max_col: int) -> list[Any]:
    from odf.table import TableCell

    width = max_col - min_col + 1
    cells: list[Any] = [None] * width
    col = 1
    for cell in row.getElementsByType(TableCell):
        repeat = int(cell.getAttribute("numbercolumnsrepeated") or "1")
        value = _ods_cell_text(cell)
        for _ in range(repeat):
            if min_col <= col <= max_col:
                cells[col - min_col] = value
            col += 1
            if col > max_col:
                break
        if col > max_col:
            break
    return cells


def _ods_sheet_dimensions(table) -> tuple[int, int]:
    from odf.table import TableRow, TableCell

    row_count = len(table.getElementsByType(TableRow))
    max_col = 1
    for row in table.getElementsByType(TableRow):
        col = 1
        for cell in row.getElementsByType(TableCell):
            repeat = int(cell.getAttribute("numbercolumnsrepeated") or "1")
            if _ods_cell_text(cell) is not None:
                max_col = max(max_col, col + repeat - 1)
            col += repeat
    return max(1, row_count), max(1, max_col)


def _trim_trailing_empty(matrix: list[list[Any]]) -> list[list[Any]]:
    if not matrix:
        return matrix
    last_col = 0
    for row in matrix:
        for idx in range(len(row) - 1, -1, -1):
            if row[idx] is not None and row[idx] != "":
                last_col = max(last_col, idx + 1)
                break
    if last_col == 0:
        return matrix
    return [row[:last_col] for row in matrix]


def _read_ods(
    path: Path,
    sheet_name: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    max_rows: int,
    _include_formulas: bool,
) -> tuple[list[list[Any]], dict[str, int], bool]:
    from odf.opendocument import load
    from odf.table import Table, TableRow

    doc = load(str(path))
    table = None
    for t in doc.spreadsheet.getElementsByType(Table):
        if t.getAttribute("name") == sheet_name:
            table = t
            break
    if table is None:
        raise ValueError(f"Sheet {sheet_name!r} not found")

    total_rows_in_window = max(0, max_row - min_row + 1)
    data_row_budget = max_rows if max_rows > 0 else total_rows_in_window
    allowed_rows = min(total_rows_in_window, 1 + data_row_budget)
    effective_max_row = min_row + allowed_rows - 1

    matrix: list[list[Any]] = []
    row_num = 0
    for row in table.getElementsByType(TableRow):
        row_num += 1
        if row_num < min_row:
            continue
        if row_num > effective_max_row:
            break
        matrix.append(_expand_ods_row(row, min_col, max_col))

    truncated = effective_max_row < max_row or (max_rows > 0 and total_rows_in_window > allowed_rows)
    dimensions = {
        "total_rows_in_range": total_rows_in_window,
        "returned_rows": len(matrix),
        "col_count": max_col - min_col + 1,
    }
    return matrix, dimensions, truncated


def _list_sheet_names(path: Path, ext: str) -> list[str]:
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    from odf.opendocument import load
    from odf.table import Table

    doc = load(str(path))
    return [t.getAttribute("name") for t in doc.spreadsheet.getElementsByType(Table)]


def read_spreadsheet(
    path: Path,
    sheet: str | None,
    max_rows: int,
    range_spec: str | None,
    include_formulas: bool,
) -> dict[str, Any]:
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported extension {ext!r}; use .xlsx or .ods")

    sheet_names = _list_sheet_names(path, ext)
    active_sheet = _resolve_sheet_name(sheet_names, sheet)

    if range_spec:
        min_row, max_row, min_col, max_col = _parse_a1_range(range_spec)
    else:
        min_row, min_col = 1, 1
        max_col = 50
        if ext == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True)
            try:
                ws = wb[active_sheet]
                max_row = ws.max_row or 1
                max_col = min(ws.max_column or 1, 50)
            finally:
                wb.close()
        else:
            from odf.opendocument import load
            from odf.table import Table

            doc = load(str(path))
            max_row, max_col = 1, 1
            for t in doc.spreadsheet.getElementsByType(Table):
                if t.getAttribute("name") == active_sheet:
                    max_row, max_col = _ods_sheet_dimensions(t)
                    break

    if ext == ".xlsx":
        matrix, dimensions, truncated = _read_xlsx(
            path,
            active_sheet,
            min_row,
            max_row,
            min_col,
            max_col,
            max_rows,
            include_formulas,
        )
    else:
        matrix, dimensions, truncated = _read_ods(
            path,
            active_sheet,
            min_row,
            max_row,
            min_col,
            max_col,
            max_rows,
            include_formulas,
        )

    matrix = _trim_trailing_empty(matrix)
    headers = matrix[0] if matrix else []
    data_rows = matrix[1:] if len(matrix) > 1 else []
    if matrix:
        dimensions["col_count"] = len(matrix[0])

    return {
        "path": str(path),
        "format": ext.lstrip("."),
        "sheet_names": sheet_names,
        "sheet": active_sheet,
        "range": range_spec or f"{_index_to_col(min_col)}{min_row}:{_index_to_col(max_col)}{max_row}",
        "headers": headers,
        "rows": data_rows,
        "truncated": truncated,
        "dimensions": dimensions,
        "include_formulas": include_formulas,
    }


def _index_to_col(index: int) -> str:
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters or "A"


def main() -> int:
    parser = argparse.ArgumentParser(description="Read spreadsheet sheet as JSON")
    parser.add_argument("path", help="Path to .xlsx or .ods file")
    parser.add_argument("--sheet", help="Sheet name or 1-based index (default: first sheet)")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=DEFAULT_MAX_ROWS,
        help=f"Max data rows after header row (default {DEFAULT_MAX_ROWS})",
    )
    parser.add_argument("--range", dest="range_spec", help="A1 range, e.g. B2:F50")
    parser.add_argument(
        "--include-formulas",
        action="store_true",
        help="Return formula strings instead of cached values (xlsx only)",
    )
    args = parser.parse_args()

    path = Path(args.path)
    if not path.is_file():
        print(json.dumps({"error": f"File not found: {path}"}))
        return 1

    try:
        result = read_spreadsheet(
            path,
            args.sheet,
            max(0, args.max_rows),
            args.range_spec,
            args.include_formulas,
        )
        print(json.dumps(result, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(json.dumps({"error": str(exc)}))
        return 1


if __name__ == "__main__":
    sys.exit(main())
