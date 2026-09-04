#!/usr/bin/env python3
"""Create or overwrite an .xlsx workbook with sheets, data, and optional charts.

Reads a workbook-definition JSON document from stdin and writes an .xlsx file.
Uses openpyxl (already a read dependency) which supports native Excel charts
(bar/column, line, pie, scatter).

Schema (stdin JSON):
{
  "path": "out.xlsx",            # required, must end with .xlsx
  "overwrite": true,             # default false; refuse to clobber existing file
  "sheets": [                    # at least one sheet
    {
      "name": "Sheet1",
      "title": "Optional title",  # written in row 1 (optional)
      "headers": ["A","B","C"],   # optional header row
      "rows": [[...],[...]],      # list of rows; each row is a list of cell values
      "column_widths": [12,20],   # optional widths in Excel width units
      "freeze_header": true,      # freeze panes at row after header (optional)
      "auto_filter": true         # enable autofilter on header range (optional)
    }
  ],
  "charts": [                    # optional charts
    {
      "type": "bar" | "line" | "pie" | "scatter",
      "sheet": "Sheet1",          # sheet containing data + where chart anchors
      "title": "Optional chart title",
      "anchor": "H2",             # A1 anchor cell; default: right of data
      "width_cm": 15,             # optional; default 15
      "height_cm": 8,             # optional; default 8
      "x_axis_title": "...",      # optional (ignored for pie)
      "y_axis_title": "...",     # optional (ignored for pie)
      "categories": "A2:A10",      # category labels range (bar/line/pie only)
      "series": [
        # bar/line/pie series:
        {"name": "Sales", "values": "B2:B10"},
        # scatter series uses x + y ranges:
        {"name": "Run", "x": "A2:A10", "y": "B2:B10"}
      ]
    }
  ]
}

Errors are emitted as JSON {"error": "..."} on stdout and a non-zero exit code.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# A1 helpers
# ---------------------------------------------------------------------------
_COL_RE = re.compile(r"^[A-Z]+$")


def _col_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letters: {letters!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _parse_a1(spec: str) -> tuple[int, int, int, int]:
    """Return min_row, max_row, min_col, max_col (1-based) for an A1 range."""
    spec = spec.strip().upper()
    m = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", spec)
    if not m:
        raise ValueError(f"Invalid range {spec!r}; use A1 notation like B2:B10")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    min_col, max_col = _col_to_index(c1), _col_to_index(c2)
    if r1 > r2 or min_col > max_col:
        raise ValueError(f"Invalid range {spec!r}; corners must form a rectangle")
    return r1, r2, min_col, max_col


# ---------------------------------------------------------------------------
# Value normalization
# ---------------------------------------------------------------------------
def _coerce_cell(value: Any) -> Any:
    """Keep native types; reject nested objects/lists (would serialize oddly)."""
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


# ---------------------------------------------------------------------------
# Sheet writing
# ---------------------------------------------------------------------------
def _write_sheet(wb: Workbook, spec: dict[str, Any], *, is_first: bool) -> int:
    name = str(spec.get("name") or "").strip()
    if not name:
        raise ValueError("Each sheet requires a non-empty 'name'")
    # openpyxl's Workbook() ships one empty default sheet named "Sheet". Reuse it
    # for the first requested sheet (rename it); create new sheets afterwards.
    if is_first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)

    next_row = 1
    title = spec.get("title")
    if title:
        ws.cell(row=next_row, column=1, value=str(title))
        ws.cell(row=next_row, column=1).font = Font(bold=True, size=14)
        next_row += 1
        # leave one blank row after a title for breathing room
        next_row += 1

    headers = spec.get("headers")
    if headers:
        if not isinstance(headers, list):
            raise ValueError("'headers' must be a list")
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=next_row, column=col, value=_coerce_cell(h))
            cell.font = Font(bold=True)
        header_row = next_row
        next_row += 1
    else:
        header_row = None

    rows = spec.get("rows") or []
    if not isinstance(rows, list):
        raise ValueError("'rows' must be a list of rows")
    max_col_used = len(headers) if headers else 0
    for row in rows:
        if not isinstance(row, list):
            raise ValueError("Each row must be a list of cell values")
        for col, val in enumerate(row, start=1):
            ws.cell(row=next_row, column=col, value=_coerce_cell(val))
        if len(row) > max_col_used:
            max_col_used = len(row)
        next_row += 1

    # Column widths
    widths = spec.get("column_widths")
    if widths:
        if not isinstance(widths, list):
            raise ValueError("'column_widths' must be a list of numbers")
        for i, w in enumerate(widths, start=1):
            if isinstance(w, (int, float)) and w > 0:
                ws.column_dimensions[get_column_letter(i)].width = float(w)

    # Freeze header
    if spec.get("freeze_header") and header_row is not None:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Auto filter on header row across used columns
    if spec.get("auto_filter") and header_row is not None and max_col_used > 0:
        last_col = get_column_letter(max(max_col_used, 1))
        last_row = max(next_row - 1, header_row)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    return max_col_used


# ---------------------------------------------------------------------------
# Chart writing
# ---------------------------------------------------------------------------
def _build_reference(ws, spec: dict[str, Any], key: str, *, allow_single: bool = False) -> Reference:
    raw = spec.get(key)
    if raw is None:
        raise ValueError(f"Chart series missing '{key}' range")
    raw = str(raw).strip()
    # Allow single-cell references by expanding to a range
    if allow_single and re.fullmatch(r"[A-Za-z]+\d+", raw):
        raw = f"{raw}:{raw}"
    r1, r2, c1, c2 = _parse_a1(raw)
    return Reference(ws, min_row=r1, max_row=r2, min_col=c1, max_col=c2)


def _add_chart(wb: Workbook, spec: dict[str, Any]) -> None:
    sheet_name = str(spec.get("sheet") or "").strip()
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Chart sheet {sheet_name!r} not found; available: {wb.sheetnames}")
    ws = wb[sheet_name]

    ctype = str(spec.get("type") or "").strip().lower()
    if ctype not in {"bar", "line", "pie", "scatter"}:
        raise ValueError(f"Unsupported chart type {ctype!r}; use bar, line, pie, or scatter")

    series_specs = spec.get("series") or []
    if not isinstance(series_specs, list) or not series_specs:
        raise ValueError("Chart requires a non-empty 'series' list")

    if ctype == "bar":
        chart = BarChart()
    elif ctype == "line":
        chart = LineChart()
    elif ctype == "pie":
        chart = PieChart()
    else:
        chart = ScatterChart()

    title = spec.get("title")
    if title:
        chart.title = str(title)
    if isinstance(spec.get("width_cm"), (int, float)) and spec["width_cm"] > 0:
        chart.width = float(spec["width_cm"])
    else:
        chart.width = 15
    if isinstance(spec.get("height_cm"), (int, float)) and spec["height_cm"] > 0:
        chart.height = float(spec["height_cm"])
    else:
        chart.height = 8

    if ctype != "scatter":
        for s in series_specs:
            if not isinstance(s, dict):
                raise ValueError("Each series must be an object")
            values = _build_reference(ws, s, "values")
            name = s.get("name")
            ser = Series(values, title=str(name) if name is not None else None)
            chart.series.append(ser)
        # Categories (optional for pie; useful for bar/line)
        cats = spec.get("categories")
        if cats:
            cat_ref = _build_reference(ws, {"categories": cats}, "categories")
            chart.set_categories(cat_ref)
        if ctype != "pie":
            if spec.get("x_axis_title"):
                chart.x_axis.title = str(spec["x_axis_title"])
            if spec.get("y_axis_title"):
                chart.y_axis.title = str(spec["y_axis_title"])
            chart.x_axis.delete = False
            chart.y_axis.delete = False
    else:
        for s in series_specs:
            if not isinstance(s, dict):
                raise ValueError("Each scatter series must be an object")
            x_ref = _build_reference(ws, s, "x")
            y_ref = _build_reference(ws, s, "y")
            name = s.get("name")
            ser = Series(y_ref, xvalues=x_ref, title=str(name) if name is not None else None)
            chart.series.append(ser)
        if spec.get("x_axis_title"):
            chart.x_axis.title = str(spec["x_axis_title"])
        if spec.get("y_axis_title"):
            chart.y_axis.title = str(spec["y_axis_title"])
        chart.x_axis.delete = False
        chart.y_axis.delete = False

    # Anchor: default to two columns right of used data, row 2.
    anchor = spec.get("anchor")
    if anchor:
        anchor = str(anchor).strip()
    else:
        max_col = 1
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
        anchor = f"{get_column_letter(max_col + 2)}2"

    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def write_spreadsheet(defn: dict[str, Any]) -> dict[str, Any]:
    out_path = Path(str(defn.get("path") or "")).expanduser()
    if out_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output path must end with .xlsx: {out_path}")
    if out_path.exists() and not defn.get("overwrite"):
        raise ValueError(
            f"File exists: {out_path}. Set overwrite=true to replace it."
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = defn.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("'sheets' must be a non-empty list")

    charts = defn.get("charts") or []

    wb = Workbook()
    sheet_names_written: list[str] = []
    for idx, sheet_spec in enumerate(sheets):
        if not isinstance(sheet_spec, dict):
            raise ValueError("Each sheet must be an object")
        _write_sheet(wb, sheet_spec, is_first=(idx == 0))
        sheet_names_written.append(wb.sheetnames[-1])

    for chart_spec in charts:
        if not isinstance(chart_spec, dict):
            raise ValueError("Each chart must be an object")
        _add_chart(wb, chart_spec)

    wb.save(out_path)

    return {
        "path": str(out_path),
        "format": "xlsx",
        "sheet_names": sheet_names_written,
        "sheets": len(sheet_names_written),
        "charts": len(charts),
        "overwrote": out_path.exists(),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Create an .xlsx workbook from a JSON definition on stdin")
    parser.add_argument("--definition", help="Path to a JSON definition file (defaults to stdin)")
    args, _unknown = parser.parse_known_args()

    try:
        if args.definition:
            raw = Path(args.definition).read_text(encoding="utf-8")
        else:
            raw = sys.stdin.read()
        defn = json.loads(raw)
        if not isinstance(defn, dict):
            raise ValueError("Definition must be a JSON object")
        result = write_spreadsheet(defn)
        print(json.dumps(result, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    sys.exit(main())